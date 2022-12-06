import os
import re
import datetime

import openpyxl
from hashlib import md5
from Const import *


def is_int(s):
    try:
        int(s)
        return True
    except ValueError:
        return False


def mkdirs(path):
    if not os.path.exists(path):
        os.makedirs(path)
    return path


def get_week(t):
    china_time = datetime.datetime.strptime(t, '%Y/%m/%dT%H:%M:%S')
    return china_time.strftime('%w')


def time_cn2utc(t, fmt='%Y%m%dT%H%M%SZ'):
    utc_time = datetime.datetime.strptime(t, '%Y/%m/%dT%H:%M:%S') - datetime.timedelta(hours=8)
    return utc_time.strftime(fmt)


def get_month(path):
    file = os.path.basename(path)
    matcher = re.search(r'(\d+)月', file)
    if matcher:
        month = matcher.group(1)
        return month
    else:
        return None


def get_name_sheet(sh, row_count, name):
    # 姓名的行列
    name_info = (0, 0)
    for y in range(2, row_count):
        for x in range(1, 5):
            if sh.cell(row=y, column=x).value == name:
                name_info = (x, y)
                break
    return name_info


def read_excel(path, name, sheet=None):
    # 打开工作簿
    workbook = openpyxl.load_workbook(path)
    if sheet is None:
        for sheet_name in workbook.sheetnames:
            if str(sheet_name).find('工时') > -1:
                continue
            # 获取表单
            sh = workbook[sheet_name]
            name_info = get_name_sheet(sh, sh.max_row, name)
            if name_info[0] > 0 and name_info[1] > 0:
                sheet = sheet_name
                break

    sh = workbook[sheet]
    row_count = sh.max_row  # 有效数据行数
    col_count = sh.max_column  # 有效数据列数

    # 姓名的行列
    name_info = get_name_sheet(sh, row_count, name)
    # 日期的行列
    date_info = (0, 0)
    for y in range(2, row_count):
        for x in range(1, 5):
            if str(sh.cell(row=y, column=x).value).find('日期') > -1:
                date_info = (x, y)
                break

    # 获取月份
    month = get_month(path)
    if not month:
        month = get_month(sh.cell(row=1, column=1).value)

    month = month if len(month) > 1 else f'0{month}'
    year = int(datetime.date.today().strftime('%Y'))
    if month == '01':  # 1月排版肯定是去年12月搞的
        year += 1
    work_info = {
        'year': str(year),
        'month': month,
        'days': []
    }
    for x in range(name_info[0] + 1, col_count):
        day = str(sh.cell(row=date_info[1], column=x).value)
        if is_int(day):
            day = day if len(day) > 1 else f'0{day}'
            work_info['days'].append({
                'day': day,
                'tag': sh.cell(row=name_info[1], column=x).value
            })
    return work_info


def main(path, name, sheet=None):
    work_list = read_excel(path, name, sheet)

    year = work_list['year']
    month = work_list['month']

    ical = f"""BEGIN:VCALENDAR
VERSION:2.0
CALSCALE:GREGORIAN
METHOD:PUBLISH
PRODID:Celetor
X-WR-CALNAME:{year}年{int(month)}月排班表"""

    for work in work_list['days']:
        title = work['tag']
        if calendar.get(title):
            start_hms = calendar[title]['start_time']
            end_hms = calendar[title]['end_time']
        else:
            continue

        date = f'{year}/{month}/{work["day"]}'
        start_time_cn = f'{date}T{start_hms}'
        start_time = time_cn2utc(start_time_cn)

        week = get_week(start_time_cn)  # 星期几
        if end_hms == '21:30:00' and (week == '5' or week == '6'):
            end_hms = '22:00:00'
        description = f'打工时间：{start_hms[:-3]}～{end_hms[:-3]}' + (
            '\\n注意：节假日22:00闭店' if end_hms == '21:30:00' else '')

        end_time_cn = f'{date}T{end_hms}'
        end_time = time_cn2utc(end_time_cn)

        ical += f'''
BEGIN:VEVENT
SUMMARY:{title}班
UID:{start_time}
DTSTART:{start_time}
DTEND:{end_time}
DESCRIPTION:{description}
END:VEVENT'''

    ical += '''
END:VCALENDAR'''
    with open(f"./{mkdirs(year + month)}/{md5(name.encode()).hexdigest()}.ics", 'w', encoding='utf-8') as f:
        f.write(ical)


if __name__ == "__main__":
    main(r'd:/XX.xlsx', 'YY')
